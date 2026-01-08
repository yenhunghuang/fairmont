"""Excel generation service for Fairmont format quotations.

完全比照惠而蒙格式 Excel 範本:
- 公司表頭 (rows 1-15)
- 欄位標題 (row 16)
- 資料列 (row 17+)
- 條款 footer

支援從 OutputFormatSkill 載入格式配置，並提供 fallback 預設值。
"""

import base64
import logging
import uuid
from datetime import datetime
from io import BytesIO
from pathlib import Path
from typing import Optional, TYPE_CHECKING

from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.drawing.spreadsheet_drawing import TwoCellAnchor, AnchorMarker
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.utils.units import pixels_to_EMU

from ..config import settings
from ..models import BOQItem, Quotation
from ..utils import ErrorCode, FileManager, raise_error
from .dimension_formatter import get_dimension_formatter_service

# Fallback 導入（當 Skill 載入失敗時使用）
from .quotation_format import (
    COLUMNS as DEFAULT_COLUMNS,
    DATA_HEADER_ROW as DEFAULT_DATA_HEADER_ROW,
    DATA_START_ROW as DEFAULT_DATA_START_ROW,
    FAIRMONT_COMPANY as DEFAULT_COMPANY,
    FAIRMONT_TERMS as DEFAULT_TERMS,
    FAIRMONT_TERMS_HEADER as DEFAULT_TERMS_HEADER,
)

if TYPE_CHECKING:
    from .skill_loader import SkillLoaderService, OutputFormatSkill

logger = logging.getLogger(__name__)


class ExcelGeneratorService:
    """Service for generating Excel quotations in Fairmont format.

    支援從 OutputFormatSkill 載入格式配置，使用 Constructor Injection。
    """

    def __init__(self, skill_loader: Optional["SkillLoaderService"] = None):
        """Initialize Excel generator service.

        Args:
            skill_loader: SkillLoaderService 實例，None 時使用全域單例
        """
        self.file_manager = FileManager(
            temp_dir=settings.temp_dir_path,
            images_dir=settings.extracted_images_dir_path,
        )
        self._skill_loader = skill_loader
        self._output_format: Optional["OutputFormatSkill"] = None
        self._format_loaded = False

    def _ensure_skill_loaded(self) -> None:
        """確保 Skill 已載入（懶載入）."""
        if self._format_loaded:
            return

        if self._skill_loader is None:
            from .skill_loader import get_skill_loader
            self._skill_loader = get_skill_loader()

        self._output_format = self._skill_loader.load_output_format_or_default("fairmont")
        self._format_loaded = True

    def _get_columns(self) -> list:
        """取得欄位定義."""
        self._ensure_skill_loaded()
        if self._output_format and self._output_format.columns:
            return [
                (col.header, col.field, col.width, col.alignment)
                for col in self._output_format.columns
            ]
        return DEFAULT_COLUMNS

    def _get_data_header_row(self) -> int:
        """取得欄位標題列."""
        self._ensure_skill_loaded()
        if self._output_format:
            return self._output_format.data_header_row
        return DEFAULT_DATA_HEADER_ROW

    def _get_data_start_row(self) -> int:
        """取得資料起始列."""
        self._ensure_skill_loaded()
        if self._output_format:
            return self._output_format.data_start_row
        return DEFAULT_DATA_START_ROW

    def _get_company_info(self) -> dict:
        """取得公司資訊."""
        self._ensure_skill_loaded()
        if self._output_format:
            return self._output_format.company.model_dump()
        return DEFAULT_COMPANY

    def _get_terms(self) -> tuple[str, list[str]]:
        """取得條款."""
        self._ensure_skill_loaded()
        if self._output_format:
            return (self._output_format.terms.header, self._output_format.terms.items)
        return (DEFAULT_TERMS_HEADER, DEFAULT_TERMS)

    def create_quotation_excel(
        self,
        quotation: Quotation,
        include_photos: bool = True,
        photo_height_cm: float = 3.0,
    ) -> str:
        """
        Create Excel file for quotation.

        Args:
            quotation: Quotation object with items
            include_photos: Whether to embed photos
            photo_height_cm: Photo height in centimeters

        Returns:
            Path to generated Excel file

        Raises:
            APIError: If generation fails
        """
        try:
            # Create workbook
            wb = Workbook()
            ws = wb.active
            ws.title = "報價單"

            # 1. Write company header (rows 1-15)
            self._write_company_header(ws, quotation)

            # 2. Write column headers (row 16)
            self._write_column_headers(ws)

            # 3. Add items (row 17+)
            self._add_items_to_sheet(ws, quotation.items, include_photos, photo_height_cm)

            # 4. Write terms footer
            footer_start_row = self._get_data_start_row() + len(quotation.items) + 2
            self._write_terms_footer(ws, footer_start_row)

            # Generate filename and save
            filename = f"quotation_{quotation.id}_{uuid.uuid4().hex[:8]}.xlsx"
            file_path = self.file_manager.temp_dir / filename

            wb.save(str(file_path))
            logger.info(f"Excel file created: {file_path}")

            return str(file_path)

        except Exception as e:
            logger.error(f"Excel generation failed: {e}")
            raise_error(
                ErrorCode.EXPORT_FAILED,
                f"Excel 生成失敗：{str(e)}",
            )

    def _write_company_header(self, ws, quotation: Quotation) -> None:
        """Write Fairmont company header section (rows 1-15)."""
        # Format date
        quote_date = quotation.created_at.strftime("%m/%d/%Y")

        # Title font
        title_font = Font(bold=True, size=18)

        # Add Fairmont logo at top-left (A1)
        self._embed_logo(ws)

        # Row 1: QUOTE title (right side)
        ws.cell(row=1, column=10, value="QUOTE").font = title_font

        # Row 5: Company address + Project Name
        company = self._get_company_info()
        ws.cell(row=5, column=2, value=company["address"])
        ws.cell(row=5, column=8, value="Project Name：")
        ws.cell(row=5, column=10, value=quotation.project_name or "")

        # Row 6: Phone
        ws.cell(row=6, column=2, value=f"p {company['phone']}")

        # Row 7: Fax + RFQ #
        ws.cell(row=7, column=2, value=f"f  {company['fax']}")
        ws.cell(row=7, column=8, value="RFQ #：")
        ws.cell(row=7, column=10, value=quotation.title or "")

        # Row 8: Website + Date
        ws.cell(row=8, column=2, value=company["website"])
        ws.cell(row=8, column=8, value="Date：")
        ws.cell(row=8, column=10, value=quote_date)

        # Row 9: Revision
        ws.cell(row=9, column=8, value="Revision：")
        ws.cell(row=9, column=10, value="1")

        # Row 11-14: Customer info + Production details
        ws.cell(row=11, column=2, value="Customer")
        ws.cell(row=11, column=8, value="Production Type：")
        ws.cell(row=11, column=10, value="BULK")

        ws.cell(row=12, column=2, value="Company：")
        ws.cell(row=12, column=8, value="Shipping Term：")
        ws.cell(row=12, column=10, value="FOB")

        ws.cell(row=13, column=2, value="ATTN：")
        ws.cell(row=13, column=8, value="Port of Loading：")
        ws.cell(row=13, column=10, value="Nantong China")

        ws.cell(row=14, column=2, value="Email：")

    def _embed_logo(self, ws) -> None:
        """Embed Fairmont logo at top-left corner (A1)."""
        # Logo file path
        logo_path = Path(__file__).parent.parent.parent.parent / "docs" / "fairmont-logo.jpg"

        if not logo_path.exists():
            logger.warning(f"Logo file not found: {logo_path}")
            return

        try:
            # Load logo image
            img = XLImage(str(logo_path))

            # Set logo size to match example template
            # Logo spans rows 1-4, columns A-B (header area)
            LOGO_WIDTH = 280
            LOGO_HEIGHT = 75

            img.width = LOGO_WIDTH
            img.height = LOGO_HEIGHT

            # Position at A1 with small offset
            from_marker = AnchorMarker(
                col=0,  # Column A (0-indexed)
                colOff=pixels_to_EMU(5),
                row=0,  # Row 1 (0-indexed)
                rowOff=pixels_to_EMU(5),
            )
            to_marker = AnchorMarker(
                col=3,  # Span to column D (0-indexed, so 3 = D)
                colOff=pixels_to_EMU(50),
                row=3,  # Span to row 4
                rowOff=pixels_to_EMU(10),
            )

            img.anchor = TwoCellAnchor(editAs='oneCell', _from=from_marker, to=to_marker)
            ws.add_image(img)

            logger.debug("Fairmont logo embedded at A1")

        except Exception as e:
            logger.warning(f"Failed to embed logo: {e}")

    def _write_column_headers(self, ws) -> None:
        """Create column header row (row 16) with formatting."""
        # Header styling
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        # Add headers at row 16
        columns = self._get_columns()
        data_header_row = self._get_data_header_row()
        for col_num, (header_text, _, excel_width, _) in enumerate(columns, 1):
            cell = ws.cell(row=data_header_row, column=col_num)
            cell.value = header_text
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
            cell.border = thin_border
            ws.column_dimensions[get_column_letter(col_num)].width = excel_width

        ws.row_dimensions[data_header_row].height = 25

    def _write_terms_footer(self, ws, start_row: int) -> None:
        """Write terms and remarks footer section."""
        terms_header, terms_items = self._get_terms()

        # Terms header
        ws.cell(row=start_row, column=1, value=terms_header)
        ws.cell(row=start_row, column=1).font = Font(bold=True)

        # Add numbered terms
        for i, term in enumerate(terms_items, 1):
            ws.cell(row=start_row + 1 + i, column=1, value=str(i))
            ws.cell(row=start_row + 1 + i, column=2, value=term)

    def _add_items_to_sheet(
        self,
        ws,
        items: list[BOQItem],
        include_photos: bool = True,
        photo_height_cm: float = 3.0,
    ) -> None:
        """Add items to worksheet starting at row 17."""
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )
        center_alignment = Alignment(horizontal="center", vertical="top", wrap_text=True)
        left_alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        right_alignment = Alignment(horizontal="right", vertical="top", wrap_text=True)

        # Calculate photo size in pixels (assuming 96 DPI)
        photo_height_px = int(photo_height_cm * 37.8)

        # Initialize dimension formatter
        dim_formatter = get_dimension_formatter_service()

        data_start_row = self._get_data_start_row()
        for idx, item in enumerate(items):
            row_num = data_start_row + idx  # Start from row 17

            # A: NO.
            cell = ws.cell(row=row_num, column=1)
            cell.value = item.no
            cell.alignment = center_alignment
            cell.border = thin_border

            # B: Item no.
            cell = ws.cell(row=row_num, column=2)
            cell.value = item.item_no
            cell.alignment = left_alignment
            cell.border = thin_border

            # C: Description
            cell = ws.cell(row=row_num, column=3)
            cell.value = item.description
            cell.alignment = left_alignment
            cell.border = thin_border

            # D: Photo (Base64)
            if include_photos and item.photo_base64:
                self._embed_base64_photo(ws, row_num, 4, item.photo_base64, photo_height_px)
            else:
                cell = ws.cell(row=row_num, column=4)
                cell.value = ""
                cell.alignment = center_alignment
                cell.border = thin_border

            # E: Dimension WxDxH (mm) - 使用 dimension_formatter 格式化
            cell = ws.cell(row=row_num, column=5)
            cell.value = dim_formatter.format_dimension(item)
            cell.alignment = left_alignment
            cell.border = thin_border

            # F: Qty
            cell = ws.cell(row=row_num, column=6)
            if item.qty is not None:
                cell.value = item.qty
                cell.number_format = "0"
            cell.alignment = center_alignment
            cell.border = thin_border

            # G: UOM
            cell = ws.cell(row=row_num, column=7)
            cell.value = item.uom or ""
            cell.alignment = center_alignment
            cell.border = thin_border

            # H: Unit Rate (USD) - 留空
            cell = ws.cell(row=row_num, column=8)
            cell.value = ""
            cell.alignment = right_alignment
            cell.border = thin_border

            # I: Amount (USD) - 留空
            cell = ws.cell(row=row_num, column=9)
            cell.value = ""
            cell.alignment = right_alignment
            cell.border = thin_border

            # J: Unit CBM
            cell = ws.cell(row=row_num, column=10)
            if item.unit_cbm is not None:
                cell.value = item.unit_cbm
                cell.number_format = "0.00"
            cell.alignment = center_alignment
            cell.border = thin_border

            # K: Total CBM (公式: =F*J)
            cell = ws.cell(row=row_num, column=11)
            if item.unit_cbm is not None and item.qty is not None:
                cell.value = f"=F{row_num}*J{row_num}"
                cell.number_format = "0.00"
            cell.alignment = center_alignment
            cell.border = thin_border

            # L: Note
            cell = ws.cell(row=row_num, column=12)
            cell.value = item.note or ""
            cell.alignment = left_alignment
            cell.border = thin_border

            # M: Location
            cell = ws.cell(row=row_num, column=13)
            cell.value = item.location or ""
            cell.alignment = left_alignment
            cell.border = thin_border

            # N: Materials Used / Specs
            cell = ws.cell(row=row_num, column=14)
            cell.value = item.materials_specs or ""
            cell.alignment = left_alignment
            cell.border = thin_border

            # O: Brand - 只有面料才顯示品牌
            cell = ws.cell(row=row_num, column=15)
            if dim_formatter.is_fabric(item):
                cell.value = item.brand or ""
            else:
                cell.value = ""
            cell.alignment = left_alignment
            cell.border = thin_border

            # Set fixed row height (85px for images)
            ws.row_dimensions[row_num].height = 85

    def _embed_base64_photo(
        self,
        ws,
        row_num: int,
        col_num: int,
        photo_base64: str,
        photo_height_px: int,
    ) -> None:
        """Embed Base64 encoded photo in worksheet with TwoCellAnchor for precise centering."""
        # Image and cell dimensions
        IMG_WIDTH = 90
        IMG_HEIGHT = 65
        ROW_HEIGHT_PX = 85
        COL_WIDTH_PX = 110  # Photo column width (~15 excel units ≈ 110px)

        # Calculate centering offsets
        h_offset = (COL_WIDTH_PX - IMG_WIDTH) // 2
        v_offset = (ROW_HEIGHT_PX - IMG_HEIGHT) // 2

        try:
            # Remove data URI prefix if present
            if photo_base64.startswith("data:"):
                photo_base64 = photo_base64.split(",", 1)[1]

            # Decode Base64 to bytes
            image_bytes = base64.b64decode(photo_base64)
            image_stream = BytesIO(image_bytes)

            # Create Excel image object from bytes
            img = XLImage(image_stream)
            img.width = IMG_WIDTH
            img.height = IMG_HEIGHT

            # Create TwoCellAnchor for precise positioning
            # AnchorMarker uses 0-indexed row/col
            from_marker = AnchorMarker(
                col=col_num - 1,
                colOff=pixels_to_EMU(h_offset),
                row=row_num - 1,
                rowOff=pixels_to_EMU(v_offset),
            )
            to_marker = AnchorMarker(
                col=col_num - 1,
                colOff=pixels_to_EMU(h_offset + IMG_WIDTH),
                row=row_num - 1,
                rowOff=pixels_to_EMU(v_offset + IMG_HEIGHT),
            )

            # Apply TwoCellAnchor - image moves/resizes with cells
            img.anchor = TwoCellAnchor(editAs='oneCell', _from=from_marker, to=to_marker)

            # Add image to worksheet (no cell reference needed with TwoCellAnchor)
            ws.add_image(img)

            logger.debug(f"Embedded centered photo at row {row_num}, col {col_num}")

        except Exception as e:
            logger.warning(f"Failed to embed Base64 photo: {e}")
            cell = ws.cell(row=row_num, column=col_num)
            cell.value = "[圖片嵌入失敗]"

    def validate_excel_file(self, file_path: str) -> bool:
        """Validate generated Excel file."""
        try:
            from openpyxl import load_workbook

            wb = load_workbook(file_path)
            ws = wb.active

            # Check headers at row 16
            columns = self._get_columns()
            data_header_row = self._get_data_header_row()
            expected_headers = [col[0] for col in columns]
            actual_headers = [ws.cell(row=data_header_row, column=i).value for i in range(1, 16)]

            if actual_headers != expected_headers:
                logger.warning(f"Header mismatch at row {data_header_row}")
                raise ValueError("Headers do not match Fairmont format")

            logger.info(f"Excel file validated: {file_path}")
            return True

        except Exception as e:
            logger.error(f"Excel validation failed: {e}")
            raise_error(
                ErrorCode.EXPORT_FAILED,
                "Excel 驗證失敗",
            )


# ============================================================
# 單例工廠
# ============================================================

_excel_generator_instance: Optional[ExcelGeneratorService] = None


def get_excel_generator(
    skill_loader: Optional["SkillLoaderService"] = None,
) -> ExcelGeneratorService:
    """取得 ExcelGeneratorService 單例實例.

    Args:
        skill_loader: 可選的 SkillLoaderService 實例

    Returns:
        ExcelGeneratorService 實例
    """
    global _excel_generator_instance
    if _excel_generator_instance is None:
        _excel_generator_instance = ExcelGeneratorService(skill_loader)
    return _excel_generator_instance
