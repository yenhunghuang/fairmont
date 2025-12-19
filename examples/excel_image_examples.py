"""
Excel 圖片嵌入範例程式碼
Python openpyxl 實用範例集

用途：
1. 快速驗證圖片嵌入功能
2. 效能測試與基準比較
3. 錯誤處理示範
4. 惠而蒙格式報價單完整範例

作者：Furniture Quotation System Team
日期：2025-12-19
"""

from openpyxl import Workbook
from openpyxl.drawing.image import Image
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side, NamedStyle
from openpyxl.utils import get_column_letter
from PIL import Image as PILImage
from PIL import ImageDraw, ImageFont
from io import BytesIO
from pathlib import Path
from typing import List, Optional, Tuple
from dataclasses import dataclass
import time
import tracemalloc
import random


# ==================== 資料模型 ====================

@dataclass
class BOQItem:
    """BOQ 項目資料模型"""
    item_no: str
    description: str
    photo: Optional[PILImage.Image]
    dimension: str
    qty: int
    uom: str
    location: str
    materials_specs: str


# ==================== 圖片生成工具 ====================

def generate_test_image(
    width: int = 200,
    height: int = 200,
    text: str = "TEST",
    bg_color: tuple = (100, 150, 200),
    text_color: tuple = (255, 255, 255)
) -> PILImage.Image:
    """
    生成測試用圖片（用於測試時無真實圖片）

    Args:
        width: 圖片寬度
        height: 圖片高度
        text: 圖片中顯示的文字
        bg_color: 背景顏色 RGB
        text_color: 文字顏色 RGB

    Returns:
        PIL Image 物件
    """
    img = PILImage.new('RGB', (width, height), color=bg_color)
    draw = ImageDraw.Draw(img)

    # 繪製文字（使用預設字型）
    try:
        # 嘗試使用較大的字型
        font = ImageFont.truetype("arial.ttf", 40)
    except:
        # 如果找不到字型，使用預設
        font = ImageFont.load_default()

    # 計算文字位置（置中）
    bbox = draw.textbbox((0, 0), text, font=font)
    text_width = bbox[2] - bbox[0]
    text_height = bbox[3] - bbox[1]
    position = ((width - text_width) // 2, (height - text_height) // 2)

    draw.text(position, text, fill=text_color, font=font)

    return img


def generate_test_dataset(
    num_items: int = 10,
    include_images: bool = True
) -> List[BOQItem]:
    """
    生成測試用 BOQ 資料集

    Args:
        num_items: 生成項目數量
        include_images: 是否包含圖片

    Returns:
        BOQ 項目列表
    """
    items = []

    # 樣本資料庫
    descriptions = ["辦公桌", "辦公椅", "會議桌", "書櫃", "文件櫃", "沙發", "茶几", "屏風"]
    locations = ["1F 辦公室", "2F 會議室", "3F 主管室", "4F 接待室"]
    materials = [
        "E1 防火板 25mm",
        "美耐板 40mm",
        "實木貼皮",
        "鋼製烤漆腳架",
        "鋁合金五爪腳",
        "網布透氣椅背"
    ]

    for i in range(num_items):
        item_no = f"{chr(65 + i // 10)}-{str(i % 10 + 1).zfill(3)}"
        description = random.choice(descriptions)

        # 隨機尺寸
        w = random.randint(600, 2000)
        h = random.randint(400, 1200)
        d = random.randint(700, 900)
        dimension = f"{w}W x {h}D x {d}H mm"

        qty = random.randint(1, 10)
        uom = random.choice(["張", "組", "套", "個"])
        location = random.choice(locations)
        materials_specs = "\n".join(random.sample(materials, k=random.randint(1, 3)))

        # 生成測試圖片
        photo = None
        if include_images:
            photo = generate_test_image(
                width=random.randint(150, 300),
                height=random.randint(150, 300),
                text=f"#{i+1}",
                bg_color=(
                    random.randint(50, 200),
                    random.randint(50, 200),
                    random.randint(50, 200)
                )
            )

        items.append(BOQItem(
            item_no=item_no,
            description=description,
            photo=photo,
            dimension=dimension,
            qty=qty,
            uom=uom,
            location=location,
            materials_specs=materials_specs
        ))

    return items


# ==================== 圖片處理工具 ====================

def resize_image_to_fit_cell(
    pil_image: PILImage.Image,
    max_width_px: int = 120,
    max_height_px: int = 120,
    maintain_aspect_ratio: bool = True
) -> Image:
    """調整圖片大小以適應儲存格"""
    original_width, original_height = pil_image.size

    if maintain_aspect_ratio:
        width_ratio = max_width_px / original_width
        height_ratio = max_height_px / original_height
        scale_ratio = min(width_ratio, height_ratio)

        new_width = int(original_width * scale_ratio)
        new_height = int(original_height * scale_ratio)
    else:
        new_width = max_width_px
        new_height = max_height_px

    resized_image = pil_image.resize(
        (new_width, new_height),
        PILImage.Resampling.LANCZOS
    )

    img_byte_arr = BytesIO()
    resized_image.save(img_byte_arr, format='PNG', optimize=True)
    img_byte_arr.seek(0)

    return Image(img_byte_arr)


def convert_image_to_png(image_source) -> Image:
    """將任意格式圖片轉換為 PNG"""
    if isinstance(image_source, PILImage.Image):
        pil_img = image_source
    elif isinstance(image_source, (str, BytesIO)):
        pil_img = PILImage.open(image_source)
    else:
        raise TypeError(f"不支援的圖片來源類型: {type(image_source)}")

    # 處理透明度
    if pil_img.mode in ('RGBA', 'LA', 'P'):
        background = PILImage.new('RGB', pil_img.size, (255, 255, 255))
        if pil_img.mode == 'P':
            pil_img = pil_img.convert('RGBA')
        background.paste(pil_img, mask=pil_img.split()[-1])
        pil_img = background
    elif pil_img.mode != 'RGB':
        pil_img = pil_img.convert('RGB')

    img_byte_arr = BytesIO()
    pil_img.save(img_byte_arr, format='PNG', optimize=True)
    img_byte_arr.seek(0)

    return Image(img_byte_arr)


# ==================== 惠而蒙格式生成器 ====================

class FairmontQuotationGenerator:
    """惠而蒙格式報價單生成器"""

    COLUMNS = [
        'Item No.',
        'Description',
        'Photo',
        'Dimension',
        'Qty',
        'UOM',
        'Location',
        'Materials Used/Specs'
    ]

    COL_ITEM_NO = 1
    COL_DESCRIPTION = 2
    COL_PHOTO = 3
    COL_DIMENSION = 4
    COL_QTY = 5
    COL_UOM = 6
    COL_LOCATION = 7
    COL_MATERIALS_SPECS = 8

    COLUMN_WIDTHS = {
        1: 10,   # Item No.
        2: 30,   # Description
        3: 20,   # Photo
        4: 15,   # Dimension
        5: 8,    # Qty
        6: 10,   # UOM
        7: 20,   # Location
        8: 35    # Materials Used/Specs
    }

    IMAGE_MAX_WIDTH = 120
    IMAGE_MAX_HEIGHT = 120
    ROW_HEIGHT_WITH_IMAGE = 90

    def __init__(self):
        self.wb = Workbook()
        self.ws = self.wb.active
        self.ws.title = "報價單"
        self._setup_styles()
        self._setup_columns()

    def _setup_styles(self):
        """設定樣式"""
        header_style = NamedStyle(name="header")
        header_style.font = Font(name='Arial', size=12, bold=True, color='FFFFFF')
        header_style.fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        header_style.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        header_style.border = Border(
            left=Side(style='thin', color='000000'),
            right=Side(style='thin', color='000000'),
            top=Side(style='thin', color='000000'),
            bottom=Side(style='thin', color='000000')
        )

        data_style = NamedStyle(name="data")
        data_style.font = Font(name='Arial', size=11)
        data_style.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        data_style.border = Border(
            left=Side(style='thin', color='CCCCCC'),
            right=Side(style='thin', color='CCCCCC'),
            top=Side(style='thin', color='CCCCCC'),
            bottom=Side(style='thin', color='CCCCCC')
        )

        text_style = NamedStyle(name="text")
        text_style.font = Font(name='Arial', size=11)
        text_style.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        text_style.border = data_style.border

        self.wb.add_named_style(header_style)
        self.wb.add_named_style(data_style)
        self.wb.add_named_style(text_style)

    def _setup_columns(self):
        """設定欄位"""
        for col_idx, width in self.COLUMN_WIDTHS.items():
            col_letter = get_column_letter(col_idx)
            self.ws.column_dimensions[col_letter].width = width

        for col_idx, col_name in enumerate(self.COLUMNS, start=1):
            cell = self.ws.cell(row=1, column=col_idx, value=col_name)
            cell.style = 'header'

        self.ws.freeze_panes = 'A2'
        self.ws.row_dimensions[1].height = 25

    def _resize_image_for_cell(self, pil_image: PILImage.Image) -> Image:
        """調整圖片大小"""
        return resize_image_to_fit_cell(
            pil_image,
            self.IMAGE_MAX_WIDTH,
            self.IMAGE_MAX_HEIGHT
        )

    def add_item(self, item: BOQItem, row: int):
        """新增 BOQ 項目"""
        self.ws.cell(row=row, column=self.COL_ITEM_NO, value=item.item_no).style = 'data'
        self.ws.cell(row=row, column=self.COL_DESCRIPTION, value=item.description).style = 'text'
        self.ws.cell(row=row, column=self.COL_DIMENSION, value=item.dimension).style = 'data'
        self.ws.cell(row=row, column=self.COL_QTY, value=item.qty).style = 'data'
        self.ws.cell(row=row, column=self.COL_UOM, value=item.uom).style = 'data'
        self.ws.cell(row=row, column=self.COL_LOCATION, value=item.location).style = 'data'
        self.ws.cell(row=row, column=self.COL_MATERIALS_SPECS, value=item.materials_specs).style = 'text'

        self.ws.row_dimensions[row].height = self.ROW_HEIGHT_WITH_IMAGE

        if item.photo:
            try:
                img = self._resize_image_for_cell(item.photo)
                cell_address = f'{get_column_letter(self.COL_PHOTO)}{row}'
                self.ws.add_image(img, cell_address)
            except Exception as e:
                print(f"警告：第 {row} 列圖片嵌入失敗 - {e}")
                self.ws.cell(row=row, column=self.COL_PHOTO, value="[圖片載入失敗]").style = 'data'
        else:
            self.ws.cell(row=row, column=self.COL_PHOTO, value="[無圖片]").style = 'data'

    def add_items_batch(self, items: List[BOQItem], start_row: int = 2):
        """批次新增 BOQ 項目"""
        for idx, item in enumerate(items):
            row = start_row + idx
            self.add_item(item, row)

            if (idx + 1) % 50 == 0:
                import gc
                gc.collect()
                print(f"已處理 {idx + 1}/{len(items)} 筆項目...")

    def save(self, output_path: str) -> str:
        """儲存報價單"""
        output_path = Path(output_path)
        output_path.parent.mkdir(parents=True, exist_ok=True)

        self.wb.save(output_path)

        file_size_mb = output_path.stat().st_size / 1024 / 1024
        print(f"報價單已儲存: {output_path} ({file_size_mb:.2f} MB)")

        return str(output_path)


# ==================== 效能測試工具 ====================

def benchmark_image_insertion(
    num_images: int = 100,
    include_images: bool = True
) -> dict:
    """
    效能基準測試

    Args:
        num_images: 測試圖片數量
        include_images: 是否包含圖片（對照測試）

    Returns:
        效能指標字典
    """
    print(f"\n開始效能測試：{num_images} 筆項目（{'有' if include_images else '無'}圖片）")
    print("=" * 60)

    tracemalloc.start()
    start_time = time.time()

    # 生成測試資料
    items = generate_test_dataset(num_images, include_images=include_images)

    # 生成 Excel
    generator = FairmontQuotationGenerator()
    generator.add_items_batch(items)

    output_file = f'output/benchmark_{num_images}_items.xlsx'
    generator.save(output_file)

    end_time = time.time()
    current, peak = tracemalloc.get_traced_memory()
    tracemalloc.stop()

    metrics = {
        'num_items': num_images,
        'include_images': include_images,
        'execution_time_seconds': end_time - start_time,
        'memory_current_mb': current / 1024 / 1024,
        'memory_peak_mb': peak / 1024 / 1024,
        'items_per_second': num_images / (end_time - start_time),
        'output_file': output_file
    }

    print(f"\n測試完成！")
    print(f"  執行時間: {metrics['execution_time_seconds']:.2f} 秒")
    print(f"  峰值記憶體: {metrics['memory_peak_mb']:.2f} MB")
    print(f"  處理速度: {metrics['items_per_second']:.2f} 項目/秒")
    print(f"  輸出檔案: {metrics['output_file']}")

    return metrics


def compare_performance():
    """比較不同場景的效能"""
    print("\n" + "=" * 60)
    print("效能比較測試")
    print("=" * 60)

    scenarios = [
        (10, True, "小型報價單 (10 項目，有圖片)"),
        (10, False, "小型報價單 (10 項目，無圖片)"),
        (50, True, "中型報價單 (50 項目，有圖片)"),
        (100, True, "大型報價單 (100 項目，有圖片)"),
    ]

    results = []

    for num_items, include_images, description in scenarios:
        print(f"\n測試場景: {description}")
        metrics = benchmark_image_insertion(num_items, include_images)
        metrics['description'] = description
        results.append(metrics)

    # 輸出比較表
    print("\n" + "=" * 60)
    print("效能比較摘要")
    print("=" * 60)
    print(f"{'場景':<30} {'時間(秒)':<12} {'記憶體(MB)':<12} {'速度(項/秒)':<12}")
    print("-" * 60)

    for r in results:
        print(f"{r['description']:<30} "
              f"{r['execution_time_seconds']:<12.2f} "
              f"{r['memory_peak_mb']:<12.2f} "
              f"{r['items_per_second']:<12.2f}")


# ==================== 範例使用案例 ====================

def example_1_basic_usage():
    """範例 1：基本使用"""
    print("\n範例 1：基本使用")
    print("=" * 60)

    # 生成 10 筆測試資料
    items = generate_test_dataset(num_items=10, include_images=True)

    # 建立報價單
    generator = FairmontQuotationGenerator()
    generator.add_items_batch(items)

    # 儲存
    output_file = generator.save('output/example_1_basic.xlsx')
    print(f"範例 1 完成！檔案: {output_file}")


def example_2_large_dataset():
    """範例 2：大量資料處理"""
    print("\n範例 2：大量資料處理")
    print("=" * 60)

    # 生成 200 筆測試資料
    items = generate_test_dataset(num_items=200, include_images=True)

    # 建立報價單
    generator = FairmontQuotationGenerator()

    # 批次處理（每 50 筆一批）
    batch_size = 50
    for i in range(0, len(items), batch_size):
        batch = items[i:i + batch_size]
        print(f"處理第 {i+1}-{min(i+batch_size, len(items))} 項...")
        generator.add_items_batch(batch, start_row=2 + i)

    # 儲存
    output_file = generator.save('output/example_2_large_dataset.xlsx')
    print(f"範例 2 完成！檔案: {output_file}")


def example_3_mixed_images():
    """範例 3：混合有無圖片的項目"""
    print("\n範例 3：混合有無圖片的項目")
    print("=" * 60)

    items = []

    # 建立 20 筆項目，隨機決定是否有圖片
    for i in range(20):
        has_image = random.random() > 0.3  # 70% 機率有圖片
        item_data = generate_test_dataset(num_items=1, include_images=has_image)[0]
        items.append(item_data)

    # 建立報價單
    generator = FairmontQuotationGenerator()
    generator.add_items_batch(items)

    # 儲存
    output_file = generator.save('output/example_3_mixed_images.xlsx')
    print(f"範例 3 完成！檔案: {output_file}")


def example_4_real_world_simulation():
    """範例 4：模擬真實世界使用場景"""
    print("\n範例 4：模擬真實世界使用場景")
    print("=" * 60)

    # 模擬從 PDF 提取的資料（不同尺寸的圖片）
    items = []

    for i in range(30):
        # 隨機決定是否有圖片（模擬 PDF 可能缺圖）
        has_image = random.random() > 0.2

        # 隨機圖片尺寸（模擬 PDF 提取的各種尺寸）
        if has_image:
            width = random.choice([100, 150, 200, 250, 300, 400])
            height = random.choice([100, 150, 200, 250, 300, 400])
            photo = generate_test_image(width, height, text=f"#{i+1}")
        else:
            photo = None

        # 隨機資料完整性（模擬 PDF 解析可能遺漏欄位）
        item = BOQItem(
            item_no=f"A-{str(i+1).zfill(3)}",
            description=random.choice(["辦公桌", "辦公椅", "會議桌", "書櫃"]),
            photo=photo,
            dimension=f"{random.randint(600, 2000)}W x {random.randint(400, 1200)}D x 750H mm" if random.random() > 0.1 else "",
            qty=random.randint(1, 10),
            uom=random.choice(["張", "組", "套"]),
            location=random.choice(["1F 辦公室", "2F 會議室"]) if random.random() > 0.1 else "",
            materials_specs=random.choice(["E1 防火板", "美耐板", "實木貼皮"]) if random.random() > 0.1 else ""
        )
        items.append(item)

    # 建立報價單
    generator = FairmontQuotationGenerator()
    generator.add_items_batch(items)

    # 儲存
    output_file = generator.save('output/example_4_real_world.xlsx')
    print(f"範例 4 完成！檔案: {output_file}")
    print(f"統計：")
    print(f"  有圖片: {sum(1 for item in items if item.photo)} 筆")
    print(f"  無圖片: {sum(1 for item in items if not item.photo)} 筆")
    print(f"  缺少尺寸: {sum(1 for item in items if not item.dimension)} 筆")
    print(f"  缺少位置: {sum(1 for item in items if not item.location)} 筆")


# ==================== 主程式 ====================

def main():
    """主程式：執行所有範例與測試"""
    print("\n" + "=" * 60)
    print("Python openpyxl 圖片嵌入範例程式")
    print("Fairmont Quotation System")
    print("=" * 60)

    # 確保輸出目錄存在
    Path('output').mkdir(exist_ok=True)

    # 執行範例
    example_1_basic_usage()
    example_2_large_dataset()
    example_3_mixed_images()
    example_4_real_world_simulation()

    # 效能測試
    compare_performance()

    print("\n" + "=" * 60)
    print("所有範例與測試完成！")
    print("請檢查 output/ 目錄查看產出的 Excel 檔案")
    print("=" * 60)


if __name__ == '__main__':
    main()
